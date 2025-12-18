from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from .forms import EnquiryForm
from .models import Enquiry
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def submit_enquiry(request):
    if request.method == 'POST':
        form = EnquiryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your enquiry has been submitted successfully! We will contact you soon.')
            return redirect('home')
        else:
            messages.error(request, 'There was an error submitting your enquiry. Please check the form and try again.')
            return redirect('home')
    return redirect('home')

def is_teacher_or_admin(user):
    return user.is_authenticated and (user.is_teacher or user.is_staff or user.is_superuser)

@login_required
@user_passes_test(is_teacher_or_admin)
def enquiry_list(request):
    # Mark all unread enquiries as read when the list is viewed
    Enquiry.objects.filter(is_read=False).update(is_read=True)
    enquiries = Enquiry.objects.all().order_by('-created_at')
    return render(request, 'enquiries/enquiry_list.html', {'enquiries': enquiries})

@login_required
@user_passes_test(is_teacher_or_admin)
def get_unread_count(request):
    count = Enquiry.objects.filter(is_read=False).count()
    return JsonResponse({'count': count})

@login_required
@user_passes_test(is_teacher_or_admin)
def enquiry_delete(request, pk):
    """
    Allow teachers/admin to delete enquiries.
    """
    enquiry = get_object_or_404(Enquiry, pk=pk)
    
    if request.method == 'POST':
        enquiry.delete()
        messages.success(request, 'Enquiry deleted successfully.')
        return redirect('enquiry_list')
    
    return render(request, 'enquiries/enquiry_confirm_delete.html', {'enquiry': enquiry})

@login_required
@user_passes_test(is_teacher_or_admin)
def export_enquiries(request):
    """
    Export all enquiries to Excel file.
    """
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enquiries"
    
    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Define headers
    headers = ['ID', 'Name', 'Email', 'Phone', 'Message', 'Submitted Date', 'Status']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Fetch all enquiries
    enquiries = Enquiry.objects.all().order_by('-created_at')
    
    # Add data rows
    for enquiry in enquiries:
        ws.append([
            enquiry.id,
            enquiry.full_name,
            enquiry.email,
            enquiry.phone_number,
            enquiry.message,
            enquiry.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'Read' if enquiry.is_read else 'Unread'
        ])
    
    # Adjust column widths
    column_widths = [10, 25, 30, 15, 50, 20, 12]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'enquiries_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
